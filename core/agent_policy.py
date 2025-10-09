"""
Agent policy layer for PBIXRay MCP Server

Provides orchestrated, guardrailed operations that an AI client (e.g., Claude)
can use as a single tool call. Keeps the server simple while centralizing
validation, safety limits, and fallbacks.
"""

import time
import os
import csv
from datetime import datetime
from typing import Any, Dict, Optional, List
from core.error_handler import ErrorHandler


class AgentPolicy:
    def __init__(self, config):
        self.config = config

    # ---- helpers ----
    def _get_preview_limit(self, max_rows: Optional[int]) -> int:
        if isinstance(max_rows, int) and max_rows > 0:
            return max_rows
        return (
            self.config.get("query.max_rows_preview", 1000)
            or self.config.get("performance.default_top_n", 1000)
            or 1000
        )

    def _get_default_perf_runs(self, runs: Optional[int]) -> int:
        if isinstance(runs, int) and runs > 0:
            return runs
        return 3

    # ---- public orchestrations ----
    def ensure_connected(self, connection_manager, connection_state, preferred_index: Optional[int] = None) -> Dict[str, Any]:
        """Ensure the server is connected to a Power BI Desktop instance."""
        if connection_state.is_connected():
            info = connection_manager.get_instance_info() or {}
            return {
                "success": True,
                "already_connected": True,
                "instance": info,
            }

        instances = connection_manager.detect_instances()
        if not instances:
            return {
                "success": False,
                "error": "No Power BI Desktop instances detected. Open a .pbix in Power BI Desktop and try again.",
                "error_type": "no_instances",
                "suggestions": [
                    "Open Power BI Desktop with a .pbix file",
                    "Wait 10–15 seconds for the model to load",
                    "Then run detection again",
                ],
            }

        index = preferred_index if preferred_index is not None else 0
        connect_result = connection_manager.connect(index)
        if not connect_result.get("success"):
            return connect_result

        connection_state.set_connection_manager(connection_manager)
        connection_state.initialize_managers()

        return {
            "success": True,
            "connected_index": index,
            "instances": instances,
            "managers_initialized": connection_state._managers_initialized,
        }

    def safe_run_dax(
        self,
        connection_state,
        query: str,
        mode: str = "auto",
        runs: Optional[int] = None,
        max_rows: Optional[int] = None,
        verbose: bool = False,
        bypass_cache: bool = False,
        include_event_counts: bool = False,
    ) -> Dict[str, Any]:
        """Validate, limit, and execute a DAX query. Optionally perform perf analysis."""
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()

        query_executor = connection_state.query_executor
        performance_analyzer = connection_state.performance_analyzer

        if not query_executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')

        # First, static analysis (syntax, complexity, patterns)
        analysis = query_executor.analyze_dax_query(query)
        if not analysis.get("success"):
            return analysis
        if analysis.get("syntax_errors"):
            return {
                "success": False,
                "error": "Query validation failed",
                "error_type": "syntax_validation_error",
                "details": analysis,
            }

        lim = self._get_preview_limit(max_rows)
        # Enforce safety limits from connection_state
        try:
            limits = connection_state.get_safety_limits()
            max_rows_cap = int(limits.get('max_rows_per_call', 10000))
            if isinstance(lim, int) and lim > 0:
                lim = min(lim, max_rows_cap)
        except Exception:
            pass
        effective_mode = (mode or "auto").lower()
        notes: List[str] = []

        # Decide mode automatically: if user supplied EVALUATE and likely table expr, do preview; else analyze if explicitly requested
        if effective_mode == "auto":
            q_upper = (query or "").strip().upper()
            # If EVALUATE and not a scalar, prefer preview
            if q_upper.startswith("EVALUATE") and "TOPN(" not in q_upper:
                do_perf = False
            else:
                do_perf = True
        else:
            do_perf = effective_mode in ("analyze",)

        if do_perf:
            # Performance analysis path
            r = self._get_default_perf_runs(runs)
            if not performance_analyzer:
                # Fallback to basic execution with a note
                basic = query_executor.validate_and_execute_dax(query, 0, bypass_cache=bypass_cache)
                basic.setdefault("notes", []).append("Performance analyzer unavailable; returned basic execution only")
                basic["success"] = basic.get("success", False)
                basic.setdefault("decision", "analyze")
                basic.setdefault("reason", "Requested performance analysis, but analyzer unavailable; returned basic execution")
                return basic
            try:
                result = performance_analyzer.analyze_query(query_executor, query, r, True, include_event_counts)
                if not result.get("success"):
                    raise RuntimeError(result.get("error") or "analysis_failed")
                result.setdefault("decision", "analyze")
                result.setdefault("reason", "Performance analysis selected to obtain FE/SE breakdown and average timings")
                return result
            except Exception as _e:
                # Graceful fallback to preview when XMLA/xEvents is blocked or errors
                q_upper = (query or "").strip().upper()
                if q_upper.startswith("EVALUATE") and "TOPN(" not in q_upper:
                    try:
                        body = (query.strip()[len("EVALUATE"):]).strip()
                        query = f"EVALUATE TOPN({lim}, {body})"
                        notes.append(f"Applied TOPN({lim}) to EVALUATE query for safety (analyze fallback)")
                    except Exception:
                        pass
                exec_result = query_executor.validate_and_execute_dax(query, lim, bypass_cache=bypass_cache)
                exec_result.setdefault("notes", []).append(
                    f"Analyzer error; returned preview instead: {str(_e)}"
                )
                exec_result.setdefault("decision", "analyze_fallback_preview")
                exec_result.setdefault("reason", "XMLA/xEvents unavailable or errored; provided successful preview with safe TOPN")
                if verbose and notes:
                    exec_result.setdefault("notes", []).extend(notes)
                return exec_result

        # Preview/standard execution path with safe TOPN
        # If query starts with EVALUATE and lacks TOPN, force safe TOPN by rewriting
        q_upper = (query or "").strip().upper()
        if q_upper.startswith("EVALUATE") and "TOPN(" not in q_upper:
            try:
                body = (query.strip()[len("EVALUATE"):]).strip()
                query = f"EVALUATE TOPN({lim}, {body})"
                notes.append(f"Applied TOPN({lim}) to EVALUATE query for safety")
            except Exception:
                pass

        exec_result = query_executor.validate_and_execute_dax(query, lim, bypass_cache=bypass_cache)
        if verbose and notes:
            exec_result.setdefault("notes", []).extend(notes)
        exec_result.setdefault("decision", "preview")
        exec_result.setdefault("reason", "Fast preview chosen to minimize latency and limit rows safely with TOPN")
        if verbose:
            # Attach lightweight analysis/suggestions for best practices visibility
            exec_result.setdefault("analysis", analysis)
        return exec_result

    def summarize_model_safely(self, connection_state) -> Dict[str, Any]:
        """Prefer lightweight summary over full exports for large models."""
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()

        model_exporter = connection_state.model_exporter
        query_executor = connection_state.query_executor
        if not model_exporter:
            return ErrorHandler.handle_manager_unavailable('model_exporter')
        if not query_executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')

        # Prefer summary
        summary = model_exporter.get_model_summary(query_executor)
        if summary.get("success"):
            return summary

        # Fallback to schema export (no expressions by default in your implementation)
        schema = {
            "success": True,
            "fallback": True,
            "schema": {
                "tables": [],
                "columns": [],
                "measures": [],
                "relationships": [],
            },
            "notes": ["Model summary unavailable; returned minimal structure"],
        }
        try:
            tables = query_executor.execute_info_query("TABLES")
            columns = query_executor.execute_info_query("COLUMNS")
            measures = query_executor.execute_info_query("MEASURES", exclude_columns=["Expression"])
            relationships = query_executor.execute_info_query("RELATIONSHIPS")
            if tables.get("success"):
                schema["schema"]["tables"] = tables.get("rows", [])
            if columns.get("success"):
                schema["schema"]["columns"] = columns.get("rows", [])
            if measures.get("success"):
                schema["schema"]["measures"] = measures.get("rows", [])
            if relationships.get("success"):
                schema["schema"]["relationships"] = relationships.get("rows", [])
        except Exception:
            pass
        return schema

    # -------- Advanced agent features --------
    def plan_query(self, intent: str, context_table: Optional[str] = None, max_rows: Optional[int] = None) -> Dict[str, Any]:
        """Return a suggested approach and safe DAX skeleton based on a simple intent string."""
        intent_lower = (intent or "").lower()
        lim = self._get_preview_limit(max_rows)
        plan = {"success": True, "intent": intent, "recommendation": "preview", "max_rows": lim}

        if any(k in intent_lower for k in ["performance", "fast", "slow", "analyze se", "analyze fe"]):
            plan["recommendation"] = "analyze"
            plan["tool"] = "safe_run_dax(mode=analyze)"
        else:
            plan["tool"] = "safe_run_dax(mode=preview)"

        if context_table:
            plan["dax"] = f"EVALUATE TOPN({lim}, '{context_table}')"
        else:
            plan["dax"] = f"EVALUATE TOPN({lim}, <table_expression_here>)"

        plan["notes"] = [
            "Use SUMMARIZECOLUMNS / SELECTCOLUMNS for shaped previews",
            "Switch to mode=analyze for SE/FE breakdown",
        ]
        return plan

    # optimize_query removed; use optimize_variants for both 2+ candidates

    def optimize_variants(
        self,
        connection_state,
        candidates: List[str],
        runs: Optional[int] = None,
    ) -> Dict[str, Any]:
        """Benchmark N DAX variants and return the fastest.

        Returns per-variant timings and a winner with minimal avg_execution_ms.
        """
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        if not candidates or len([c for c in candidates if (c or "").strip()]) < 2:
            return {"success": False, "error": "Provide at least two non-empty candidates", "error_type": "invalid_input"}

        query_executor = connection_state.query_executor
        perf = connection_state.performance_analyzer
        r = self._get_default_perf_runs(runs)

        def run_bench(q: str) -> Dict[str, Any]:
            q = q or ""
            if perf:
                res = perf.analyze_query(query_executor, q, r, True)
                summary = res.get("summary") or {}
                avg_ms = summary.get("avg_execution_ms") or 0
                return {"success": bool(res.get("success")), "execution_time_ms": avg_ms, "raw": res}
            else:
                res = query_executor.validate_and_execute_dax(q, 0)
                return {"success": bool(res.get("success")), "execution_time_ms": res.get("execution_time_ms", 0), "raw": res}

        results: List[Dict[str, Any]] = []
        for idx, cand in enumerate(candidates):
            results.append({"candidate": idx, **run_bench(cand)})

        # Choose the minimal positive time; if all zero/False, mark no_winner
        valid = [r for r in results if r.get("success") and r.get("execution_time_ms", 0) >= 0]
        if not valid:
            return {"success": False, "error": "All candidates failed", "results": results}
        winner = min(valid, key=lambda r: r.get("execution_time_ms", float("inf")))
        return {"success": True, "runs": r, "results": results, "winner": winner, "decision": "optimize_variants", "reason": "Selected variant with minimum average execution time across runs"}

    def decide_and_run(
        self,
        connection_manager,
        connection_state,
        goal: str,
        query: Optional[str] = None,
        candidates: Optional[List[str]] = None,
        runs: Optional[int] = None,
        max_rows: Optional[int] = None,
        verbose: bool = False,
    ) -> Dict[str, Any]:
        """Performance-first decision orchestrator.

        - Ensures connection
        - If multiple candidates → benchmark and return the winner
        - Else if goal mentions performance → analyze query
        - Else → fast preview with safe TOPN
        """
        actions: List[Dict[str, Any]] = []
        ensured = self.ensure_connected(connection_manager, connection_state)
        actions.append({"action": "ensure_connected", "result": ensured})
        if not ensured.get("success"):
            return {"success": False, "phase": "ensure_connected", "actions": actions, "final": ensured}

        text = (goal or "").lower()
        nonempty_candidates = [c for c in (candidates or []) if (c or "").strip()]

        if len(nonempty_candidates) >= 2:
            bench = self.optimize_variants(connection_state, nonempty_candidates, runs)
            actions.append({"action": "optimize_variants", "result": bench})
            return {"success": bench.get("success", False), "decision": "optimize_variants", "reason": "Multiple candidates provided; benchmarking to pick the fastest", "actions": actions, "final": bench}

        if any(k in text for k in ["analyze", "performance", "perf", "se/fe", "storage engine", "formula engine"]):
            if not query:
                return {"success": False, "error": "No query provided for performance analysis", "phase": "input_validation", "actions": actions}
            res = self.safe_run_dax(connection_state, query, mode="analyze", runs=runs, max_rows=max_rows, verbose=verbose)
            actions.append({"action": "safe_run_dax(analyze)", "result": res})
            return {"success": res.get("success", False), "decision": "analyze", "reason": "Goal indicates performance focus; running analyzer for FE/SE breakdown", "actions": actions, "final": res}

        # Default to preview for speed
        res = self.safe_run_dax(connection_state, query or "", mode="preview", runs=runs, max_rows=max_rows, verbose=verbose)
        actions.append({"action": "safe_run_dax(preview)", "result": res})
        return {"success": res.get("success", False), "decision": "preview", "reason": "No performance focus or multiple candidates; using fast safe preview for minimal latency", "actions": actions, "final": res}

    def agent_health(self, connection_manager, connection_state) -> Dict[str, Any]:
        """Consolidated health: server info, connection status, and a quick summary."""
        info = {
            "connected": connection_state.is_connected(),
            "managers_initialized": getattr(connection_state, "_managers_initialized", False),
        }
        if not info["connected"]:
            return {"success": True, "summary": info}
        # If connected, attempt a small summary
        model_summary = self.summarize_model_safely(connection_state)
        return {"success": True, "summary": info, "model": model_summary}

    def generate_docs_safe(self, connection_state) -> Dict[str, Any]:
        """Generate documentation, preferring safe/lightweight operations for large models."""
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        exporter = connection_state.model_exporter
        executor = connection_state.query_executor
        if not exporter:
            return ErrorHandler.handle_manager_unavailable('model_exporter')
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')

        # Use get_model_summary first to get scale hints
        summary = exporter.get_model_summary(executor)
        notes: List[str] = []
        if not summary.get("success"):
            notes.append("Model summary unavailable; proceeding to basic documentation")
            return exporter.generate_documentation(executor)

        # If model scale is high, prefer lightweight docs
        counts = summary.get("counts") or {}
        # Prefer named counts if provided by exporter; otherwise fallback heuristics
        measures_count = counts.get("measures", counts.get("measure_count", 0))
        tables_count = counts.get("tables", counts.get("table_count", 0))
        columns_count = counts.get("columns", counts.get("column_count", 0))
        relationships_count = counts.get("relationships", counts.get("relationship_count", 0))

        is_large = (
            measures_count > 2000 or
            tables_count > 200 or
            columns_count > 10000 or
            relationships_count > 5000
        )
        if is_large:
            notes.append("Large model detected; generating lightweight documentation")
        doc = exporter.generate_documentation(executor)
        if notes:
            doc.setdefault("notes", []).extend(notes)
        return doc

    # -------- New utilities & orchestrations --------
    def warm_query_cache(self, connection_state, queries: List[str], runs: Optional[int] = 1, clear_cache: bool = False) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        executor = connection_state.query_executor
        perf = connection_state.performance_analyzer
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')

        r = max(1, int(runs or 1))
        stats: List[Dict[str, Any]] = []
        if clear_cache:
            try:
                executor.flush_cache()
                if perf:
                    # Best-effort engine cache clear
                    perf._clear_cache(executor)  # type: ignore[attr-defined]
            except Exception:
                pass
        for q in queries or []:
            times = []
            ok = True
            for _ in range(r):
                start = time.time()
                res = executor.validate_and_execute_dax(q, 0, bypass_cache=False)
                ok = ok and bool(res.get('success'))
                times.append((time.time() - start) * 1000)
            stats.append({'query': q, 'success': ok, 'runs': r, 'avg_ms': round(sum(times)/len(times), 2) if times else 0})
        return {'success': True, 'warmed': len(stats), 'results': stats}

    def analyze_queries_batch(self, connection_state, queries: List[str], runs: Optional[int] = 3, clear_cache: bool = True, include_event_counts: bool = False) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        executor = connection_state.query_executor
        perf = connection_state.performance_analyzer
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')
        r = self._get_default_perf_runs(runs)
        results: List[Dict[str, Any]] = []
        for q in queries or []:
            if perf:
                results.append(perf.analyze_query(executor, q, r, bool(clear_cache), include_event_counts))
            else:
                # basic timing fallback
                start = time.time()
                res = executor.validate_and_execute_dax(q, 0)
                ms = (time.time() - start) * 1000
                results.append({'success': res.get('success', False), 'query': q, 'summary': {'avg_execution_ms': round(ms, 2), 'note': 'analyzer unavailable'}})
        return {'success': True, 'runs': r, 'items': results}

    def set_cache_policy(self, connection_state, ttl_seconds: Optional[int] = None) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        executor = connection_state.query_executor
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')
        changed = {}
        if isinstance(ttl_seconds, int) and ttl_seconds >= 0:
            executor.cache_ttl_seconds = ttl_seconds
            executor.flush_cache()
            changed['cache_ttl_seconds'] = ttl_seconds
        return {'success': True, 'changed': changed, 'current': {'cache_ttl_seconds': executor.cache_ttl_seconds}}

    def profile_columns(self, connection_state, table: str, columns: Optional[List[str]] = None) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        executor = connection_state.query_executor
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')
        # Build per-column profiling queries
        cols = columns or []
        if not cols:
            # fetch all columns for table
            info = executor.execute_info_query('COLUMNS', table_name=table)
            if not info.get('success'):
                return info
            cols = [c.get('Name') for c in info.get('rows', []) if c.get('Name')]
        results: List[Dict[str, Any]] = []
        for col in cols[:200]:  # safety cap
            q = (
                f"EVALUATE ROW(\"Min\", MIN('{table}'[{col}]), "
                f"\"Max\", MAX('{table}'[{col}]), "
                f"\"Distinct\", DISTINCTCOUNT('{table}'[{col}]), "
                f"\"Nulls\", COUNTBLANK('{table}'[{col}]))"
            )
            r = executor.validate_and_execute_dax(q, 0)
            results.append({'column': col, 'success': r.get('success', False), 'stats': (r.get('rows') or [{}])[0] if r.get('rows') else {}})
        return {'success': True, 'table': table, 'columns': len(results), 'results': results}

    def get_value_distribution(self, connection_state, table: str, column: str, top_n: int = 50) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        executor = connection_state.query_executor
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')
        q = (
            f"EVALUATE TOPN({int(top_n)}, "
            f"SUMMARIZECOLUMNS('{table}'[{column}], \"Count\", COUNTROWS('{table}')), [Count], DESC)"
        )
        return executor.validate_and_execute_dax(q, 0)

    def validate_best_practices(self, connection_state) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        executor = connection_state.query_executor
        validator = connection_state.model_validator
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')
        issues: List[Dict[str, Any]] = []
        # Include integrity issues if validator available
        if validator:
            integrity = validator.validate_model_integrity()
            if integrity.get('success'):
                for i in integrity.get('issues', []):
                    issues.append(i)
        # Simple naming checks
        tables = executor.execute_info_query('TABLES')
        if tables.get('success'):
            for t in tables.get('rows', []):
                name = t.get('Name') or ''
                if name != name.strip():
                    issues.append({'type': 'naming', 'severity': 'low', 'object': f"Table:{name}", 'description': 'Leading/trailing spaces in table name'})
        measures = executor.execute_info_query('MEASURES')
        if measures.get('success'):
            for m in measures.get('rows', []):
                name = m.get('Name') or ''
                if ' ' in name and name.strip().endswith(')'):
                    pass
                # Example heuristic: discourage very short names
                if len(name.strip()) < 2:
                    issues.append({'type': 'naming', 'severity': 'low', 'object': f"Measure:{name}", 'description': 'Very short measure name'})
        return {'success': True, 'issues': issues, 'total_issues': len(issues)}

    def generate_documentation_profiled(self, connection_state, format: str = 'markdown', include_examples: bool = False) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        exporter = connection_state.model_exporter
        executor = connection_state.query_executor
        if not exporter:
            return ErrorHandler.handle_manager_unavailable('model_exporter')
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')
        doc = exporter.generate_documentation(executor)
        doc.setdefault('format', format)
        doc.setdefault('include_examples', include_examples)
        return doc

    def export_compact_schema(self, connection_state, include_hidden: bool = True) -> Dict[str, Any]:
        """Export expression-free schema (tables/columns/measures/relationships) via exporter.

        Useful for quick, diff-friendly structure snapshots without large DAX expressions.
        """
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        exporter = connection_state.model_exporter
        if not exporter:
            return ErrorHandler.handle_manager_unavailable('model_exporter')
        try:
            return exporter.export_compact_schema(bool(include_hidden))
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def export_compact_schema_xlsx(self, connection_state, include_hidden: bool = True, output_dir: Optional[str] = None) -> Dict[str, Any]:
        """Export compact schema to an Excel workbook with separate sheets for tables, columns, measures, relationships."""
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        exporter = connection_state.model_exporter
        if not exporter:
            return ErrorHandler.handle_manager_unavailable('model_exporter')
        try:
            data = exporter.export_compact_schema(bool(include_hidden))
            if not data.get('success'):
                return data
            schema = (data or {}).get('schema') or {}
            # Prepare output directory
            root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            if output_dir and isinstance(output_dir, str) and output_dir.strip():
                out_dir = output_dir.strip()
            else:
                out_dir = os.path.join(root_dir, 'exports')
            try:
                os.makedirs(out_dir, exist_ok=True)
            except Exception:
                out_dir = os.path.join(root_dir, 'exports')
                os.makedirs(out_dir, exist_ok=True)
            # Build workbook
            try:
                from openpyxl import Workbook  # type: ignore
            except Exception as e:
                return {'success': False, 'error': f'openpyxl not available: {e}'}
            wb = Workbook()
            ws_tables = getattr(wb, 'active', None)
            if ws_tables is None:
                raise RuntimeError('openpyxl workbook has no active worksheet')
            try:
                ws_tables.title = 'Tables'
            except Exception:
                pass
            ws_columns = wb.create_sheet('Columns')
            ws_measures = wb.create_sheet('Measures')
            ws_relationships = wb.create_sheet('Relationships')
            # Tables sheet
            ws_tables.append(['Table', 'Hidden'])
            for t in schema.get('tables', []) or []:
                ws_tables.append([t.get('name'), t.get('hidden')])
            # Columns sheet
            ws_columns.append(['Table', 'Column', 'Data Type', 'Hidden', 'Summarize By'])
            for t in schema.get('tables', []) or []:
                for c in t.get('columns', []) or []:
                    ws_columns.append([t.get('name'), c.get('name'), c.get('data_type'), c.get('hidden'), c.get('summarize_by')])
            # Measures sheet
            ws_measures.append(['Table', 'Measure', 'Format String', 'Display Folder', 'Hidden'])
            for t in schema.get('tables', []) or []:
                for m in t.get('measures', []) or []:
                    ws_measures.append([t.get('name'), m.get('name'), m.get('format_string'), m.get('display_folder'), m.get('hidden')])
            # Relationships sheet
            ws_relationships.append(['From Table', 'From Column', 'To Table', 'To Column', 'Active', 'Direction', 'Cardinality'])
            for r in schema.get('relationships', []) or []:
                f = r.get('from') or {}
                to = r.get('to') or {}
                ws_relationships.append([f.get('table'), f.get('column'), to.get('table'), to.get('column'), r.get('active'), r.get('direction'), r.get('cardinality')])
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            path = os.path.join(out_dir, f"compact_schema_{ts}.xlsx")
            wb.save(path)
            return {'success': True, 'format': 'xlsx', 'file': path, 'statistics': data.get('statistics'), 'database_name': data.get('database_name')}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def relationship_overview(self, connection_state) -> Dict[str, Any]:
        """Return relationships list plus optional cardinality checks.

        This prevents brittle DMV/DAX attempts from the client and offers
        a stable, typed response for relationship exploration.
        """
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        executor = connection_state.query_executor
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')
        rels = executor.execute_info_query('RELATIONSHIPS')
        result: Dict[str, Any] = {'success': bool(rels.get('success')), 'relationships': rels.get('rows', []), 'count': len(rels.get('rows', []) if rels.get('success') else [])}
        # Optionally attach issue analysis when available
        perf_opt = getattr(connection_state, 'performance_optimizer', None)
        try:
            if perf_opt:
                analysis = perf_opt.analyze_relationship_cardinality()
                result['analysis'] = analysis
        except Exception:
            # Non-fatal; just omit analysis
            pass
        return result

    def create_model_changelog(self, connection_state, reference_tmsl) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        exporter = connection_state.model_exporter
        if not exporter:
            return ErrorHandler.handle_manager_unavailable('model_exporter')
        diff = exporter.compare_models(reference_tmsl)
        summary: Dict[str, Any] = {'notes': 'Summary generated by server'}
        if isinstance(diff, dict):
            keys_list = [str(k) for k in diff.keys()]
            summary['keys'] = keys_list
        return {'success': True, 'diff': diff, 'summary': summary}

    def get_measure_impact(self, connection_state, table: str, measure: str, depth: Optional[int] = 3) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        dep = connection_state.dependency_analyzer
        if not dep:
            return ErrorHandler.handle_manager_unavailable('dependency_analyzer')
        return dep.analyze_measure_dependencies(table, measure, depth or 3)

    # removed: get_column_usage_heatmap

    def auto_document(self, connection_manager, connection_state, profile: str = 'light', include_lineage: bool = False) -> Dict[str, Any]:
        actions: List[Dict[str, Any]] = []
        ensured = self.ensure_connected(connection_manager, connection_state)
        actions.append({'action': 'ensure_connected', 'result': ensured})
        if not ensured.get('success'):
            return {'success': False, 'phase': 'ensure_connected', 'actions': actions, 'final': ensured}
        summary = self.summarize_model_safely(connection_state)
        actions.append({'action': 'summarize_model_safely', 'result': summary})
        docs = self.generate_docs_safe(connection_state)
        actions.append({'action': 'generate_docs_safe', 'result': docs})
        return {'success': docs.get('success', False), 'actions': actions, 'final': docs}

    def auto_analyze_or_preview(self, connection_manager, connection_state, query: str, runs: Optional[int] = None, max_rows: Optional[int] = None, priority: str = 'depth') -> Dict[str, Any]:
        # priority: 'speed' -> preview, 'depth' -> analyze
        mode = 'preview' if (priority or 'depth').lower() == 'speed' else 'analyze'
        return self.safe_run_dax(connection_state, query, mode=mode, runs=runs, max_rows=max_rows)

    def apply_recommended_fixes(self, connection_state, actions: List[str]) -> Dict[str, Any]:
        # Currently returns a plan; actual mutations may require TOM APIs not exposed here.
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        plan = []
        for a in actions or []:
            if a == 'hide_keys':
                plan.append({'action': a, 'status': 'suggestion', 'note': 'Hide IsKey columns in report view'})
            elif a == 'fix_summarization':
                plan.append({'action': a, 'status': 'suggestion', 'note': 'Set default summarization for numeric columns appropriately'})
            elif a == 'organize_folders':
                plan.append({'action': a, 'status': 'suggestion', 'note': 'Group measures into display folders by subject'})
            else:
                plan.append({'action': a, 'status': 'unknown'})
        return {'success': True, 'plan': plan}

    def set_performance_trace(self, connection_state, enabled: bool) -> Dict[str, Any]:
        perf = connection_state.performance_analyzer
        if not perf:
            return ErrorHandler.handle_manager_unavailable('performance_analyzer')
        if enabled:
            ok = perf.connect_amo()
            if not ok:
                return {'success': False, 'error': 'AMO not available'}
            started = perf.start_session_trace()
            return {'success': bool(started), 'trace_active': perf.trace_active}
        else:
            perf.stop_session_trace()
            return {'success': True, 'trace_active': False}

    def format_dax(self, expression: str) -> Dict[str, Any]:
        # Minimal, non-destructive formatting: trim and collapse excessive spaces
        if expression is None:
            return {'success': False, 'error': 'No expression provided'}
        text = expression.strip()
        while '  ' in text:
            text = text.replace('  ', ' ')
        return {'success': True, 'formatted': text}

    def export_model_overview(self, connection_state, format: str = 'json', include_counts: bool = True) -> Dict[str, Any]:
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        exporter = connection_state.model_exporter
        executor = connection_state.query_executor
        if not exporter:
            return ErrorHandler.handle_manager_unavailable('model_exporter')
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')
        summary = exporter.get_model_summary(executor)
        if not summary.get('success'):
            return summary
        return {'success': True, 'format': format, 'overview': summary}

    def export_flat_schema_samples(self, connection_state, format: str = 'csv', rows: int = 3, extras: Optional[List[str]] = None, output_dir: Optional[str] = None) -> Dict[str, Any]:
        """
        Export a flat list of all tables and columns with up to N sample values per column.

        Columns: Table Name, Column Name, Data Type Column, Sample Row 1 Value, Sample Row 2 Value, Sample Row 3 Value
        """
        if not connection_state.is_connected():
            return ErrorHandler.handle_not_connected()
        executor = connection_state.query_executor
        if not executor:
            return ErrorHandler.handle_manager_unavailable('query_executor')

        # Sanitize inputs
        try:
            n = int(rows)
        except Exception:
            n = 3
        n = max(1, min(n, 10))  # safety cap

        # Normalize and validate extras list (extensible but safe by default)
        extras = extras or []
        try:
            extras = [str(x).strip() for x in extras if str(x).strip()]
        except Exception:
            extras = []
        # Cheap, model-metadata-based extras we can pull from DMV rows without extra queries
        # Keep Description out as requested; can be re-added later if needed
        ALLOWED_DMV_EXTRAS = {
            'IsHidden', 'IsNullable', 'IsKey', 'SummarizeBy'
        }
        extras_dmv = [e for e in extras if e in ALLOWED_DMV_EXTRAS]

        # Helpers for DAX quoting
        def dq_table(name: str) -> str:
            return f"'{str(name).replace("'", "''")}'"

        def dq_col(name: str) -> str:
            return f"[{str(name).replace(']', ']]')}]"

        try:
            # 1) Get all columns and table map for robust name resolution
            # Exclude hidden technical RowNumber-style columns and hidden columns by default
            cols_res = executor.execute_info_query("COLUMNS", filter_expr='[IsHidden] = FALSE')
            if not cols_res.get('success'):
                return cols_res
            cols = cols_res.get('rows', []) or []

            # Attempt to get authoritative data types via TOM when available
            tom_types_map: Dict[str, Dict[str, Any]] = {}
            try:
                ttypes = executor.get_column_datatypes_tom()  # type: ignore[attr-defined]
                if isinstance(ttypes, dict) and ttypes.get('success'):
                    tom_types_map = ttypes.get('map', {}) or {}
            except Exception:
                tom_types_map = {}

            # Build a map of TableID -> Name using INFO.TABLES()
            table_map: Dict[str, str] = {}
            try:
                t_res = executor.execute_info_query("TABLES")
                if t_res.get('success'):
                    for tr in t_res.get('rows', []) or []:
                        tid = tr.get('ID') or tr.get('[ID]') or tr.get('TableID') or tr.get('[TableID]')
                        tname = tr.get('Name') or tr.get('[Name]')
                        if tid is not None and tname:
                            table_map[str(tid)] = str(tname)
            except Exception:
                pass

            # Group by table and capture data type per column
            by_table: Dict[str, List[Dict[str, Any]]] = {}
            for r in cols:
                # Resolve table name using direct fields or TableID mapping
                t = (
                    r.get('Table')
                    or r.get('TableName')
                    or r.get('TABLE_NAME')
                    or r.get('[Table]')
                    or r.get('[TableName]')
                    or r.get('[TABLE_NAME]')
                )
                if not t:
                    tid = r.get('TableID') or r.get('[TableID]') or r.get('TABLE_ID') or r.get('[TABLE_ID]')
                    if tid is not None:
                        # Prefer local table_map; only fall back to executor mapping if missing
                        t = table_map.get(str(tid))
                        if not t:
                            try:
                                t_name = executor._get_table_name_from_id(tid)  # type: ignore[attr-defined]
                                t = t_name
                            except Exception:
                                t = None
                # Determine display name (for DAX) and technical name (do not use technical for DAX references)
                display_name = (
                    r.get('Name')
                    or r.get('[Name]')
                    or r.get('COLUMN_NAME')
                    or r.get('[COLUMN_NAME]')
                )
                technical_name = (
                    r.get('ExplicitName')
                    or r.get('[ExplicitName]')
                    or r.get('InferredName')
                    or r.get('[InferredName]')
                )
                if not t:
                    continue
                # DataType may appear as text or numeric; capture broadly and normalize later
                # Prefer TOM-derived type when available
                dt = None
                try:
                    if t in tom_types_map:
                        disp_for_type = (display_name or technical_name)
                        if disp_for_type and disp_for_type in tom_types_map[t]:
                            dt = tom_types_map[t][disp_for_type]
                except Exception:
                    dt = None
                # Fallback to DMV-provided datatype
                if dt is None:
                    dt = (
                        r.get('DataType')
                        or r.get('[DataType]')
                        or r.get('ExplicitDataType')
                        or r.get('[ExplicitDataType]')
                    )
                # Exclude technical RowNumber columns
                name_check = str(display_name or technical_name or '')
                if name_check.startswith('RowNumber-'):
                    continue
                # Skip binary columns from sampling
                try:
                    _dt_lower = str(dt).lower() if dt is not None else ''
                except Exception:
                    _dt_lower = ''
                sampleable_type = ("binary" not in _dt_lower)
                sampleable = bool(
                    (display_name is not None and str(display_name).strip() != '') or
                    (technical_name is not None and str(technical_name).strip() != '')
                ) and sampleable_type
                name_out = str(display_name or technical_name or '')
                if not name_out:
                    # If we truly don't have any name, skip this row
                    continue
                by_table.setdefault(str(t), []).append({
                    'name': name_out,
                    'datatype': dt,
                    'sampleable': sampleable,
                    'display': display_name,
                    'technical': technical_name,
                    'dmv': r,  # preserve original DMV row for extras
                })

            # Normalize datatype to friendly names favoring TOM text values
            TEXT_DTYPE_MAP = {
                'int64': 'Integer',
                'wholenumber': 'Integer',
                'integer': 'Integer',
                'double': 'Double',
                'decimal': 'Decimal',
                'currency': 'Currency',
                'boolean': 'Boolean',
                'string': 'String',
                'datetime': 'DateTime',
                'date': 'Date',
                'time': 'Time',
                'binary': 'Binary',
                'variant': 'Variant',
            }
            def map_dtype(dt_val: Any) -> Any:
                try:
                    if dt_val is None:
                        return None
                    if isinstance(dt_val, (int, float)):
                        # Avoid guessing; surface the numeric code if engines emit one
                        return int(dt_val)
                    if isinstance(dt_val, str):
                        key = dt_val.strip()
                        low = key.lower()
                        # Some enums stringify as 'DataType.Int64' -> take last token
                        if '.' in low:
                            low = low.split('.')[-1]
                        return TEXT_DTYPE_MAP.get(low, key)
                    return dt_val
                except Exception:
                    return dt_val

            # 2) For each table, fetch top N rows across all columns once
            result_rows: List[List[Any]] = []
            total_tables = 0
            def _norm_key(s: Optional[str]) -> str:
                raw = (s or '').strip()
                low = raw.lower()
                # If in the form Table[Column], prefer the bracketed token
                if '[' in low and ']' in low:
                    lbi = low.rfind('[')
                    rbi = low.rfind(']')
                    if rbi > lbi >= 0:
                        low = low[lbi + 1:rbi]
                # If still qualified with dots, take the last segment
                if '.' in low:
                    low = low.split('.')[-1]
                # Strip quotes, brackets, spaces, underscores
                low = low.replace("'", "").replace("[", "").replace("]", "")
                low = low.replace(" ", "").replace("_", "")
                return low

            for table_name, col_list in by_table.items():
                if not col_list:
                    continue
                total_tables += 1
                # 3) Emit one row per column with up to n sample values
                for col in col_list:
                    col_name = col['name']
                    # Resolve data type: prefer TOM map with normalized lookup
                    dt_raw = col.get('datatype')
                    dt_resolved = None
                    try:
                        if table_name in tom_types_map:
                            # exact match
                            disp = str(col.get('display') or col_name)
                            if disp in tom_types_map[table_name]:
                                dt_resolved = tom_types_map[table_name][disp]
                            else:
                                # try case-insensitive/normalized match
                                norm_target = _norm_key(disp)
                                for k, v in tom_types_map[table_name].items():
                                    if _norm_key(str(k)) == norm_target:
                                        dt_resolved = v
                                        break
                            if dt_resolved is None and col.get('technical'):
                                tech = str(col.get('technical'))
                                if tech in tom_types_map[table_name]:
                                    dt_resolved = tom_types_map[table_name][tech]
                                else:
                                    norm_target = _norm_key(tech)
                                    for k, v in tom_types_map[table_name].items():
                                        if _norm_key(str(k)) == norm_target:
                                            dt_resolved = v
                                            break
                    except Exception:
                        dt_resolved = None
                    dt = map_dtype(dt_resolved if dt_resolved is not None else dt_raw)

                    # Per-column sampling using a fixed alias to guarantee value extraction
                    values: List[Any] = []
                    if col.get('sampleable'):
                        base_name = str(col.get('display') or col_name)
                        # Build alias table once for robust filtering on [V]
                        val_ref = f"{dq_table(table_name)}{dq_col(base_name)}"
                        alias_tbl = f"SELECTCOLUMNS({dq_table(table_name)}, \"V\", {val_ref})"
                        # Strict non-blank, non-empty/whitespace filter
                        qcol = (
                            f"EVALUATE TOPN({n}, "
                            f"FILTER({alias_tbl}, NOT(ISBLANK([V])) && COALESCE(LEN(TRIM([V] & \"\")), 0) > 0))"
                        )
                        data_col = executor.validate_and_execute_dax(qcol, n)
                        # If failed or returned no rows, try simpler filter (exclude only BLANK())
                        rows_col = data_col.get('rows', []) if data_col.get('success') else []
                        if (not data_col.get('success') or not rows_col) and col.get('technical'):
                            # Retry using technical name if display name failed
                            tech = str(col.get('technical'))
                            val_ref = f"{dq_table(table_name)}{dq_col(tech)}"
                            alias_tbl = f"SELECTCOLUMNS({dq_table(table_name)}, \"V\", {val_ref})"
                            qcol = (
                                f"EVALUATE TOPN({n}, "
                                f"FILTER({alias_tbl}, NOT(ISBLANK([V])) && COALESCE(LEN(TRIM([V] & \"\")), 0) > 0))"
                            )
                            data_col = executor.validate_and_execute_dax(qcol, n)
                            rows_col = data_col.get('rows', []) if data_col.get('success') else []
                        if data_col.get('success') and not rows_col:
                            # Relax: exclude only BLANK() to recover values
                            qcol_relax = f"EVALUATE TOPN({n}, FILTER({alias_tbl}, NOT(ISBLANK([V]))))"
                            data_col = executor.validate_and_execute_dax(qcol_relax, n)
                            rows_col = data_col.get('rows', []) if data_col.get('success') else []
                        if data_col.get('success') and not rows_col:
                            # Final fallback: unfiltered alias table to at least return something if available
                            qcol_any = f"EVALUATE TOPN({n}, {alias_tbl})"
                            data_col = executor.validate_and_execute_dax(qcol_any, n)
                            rows_col = data_col.get('rows', []) if data_col.get('success') else []
                        rows_col = data_col.get('rows', []) if data_col.get('success') else []
                        for i in range(n):
                            if i < len(rows_col) and isinstance(rows_col[i], dict):
                                rowi = rows_col[i]
                                # Prefer alias 'V', else take the first available value
                                v = rowi.get('V')
                                if v is None and rowi:
                                    try:
                                        v = next(iter(rowi.values()))
                                    except Exception:
                                        v = None
                                values.append(v)
                            else:
                                values.append(None)
                    else:
                        values = [None, None, None]
                    # Compute requested extras values from DMV row (cheap extras only)
                    extra_vals: List[Any] = []
                    if extras_dmv:
                        dmv_row = col.get('dmv') or {}
                        for ex in extras_dmv:
                            # Use common variants of field names from TMSCHEMA/CSDL
                            if ex == 'Description':
                                extra_vals.append(dmv_row.get('Description') or dmv_row.get('[Description]'))
                            elif ex == 'IsHidden':
                                extra_vals.append(dmv_row.get('IsHidden') if 'IsHidden' in dmv_row else dmv_row.get('[IsHidden]'))
                            elif ex == 'IsNullable':
                                extra_vals.append(dmv_row.get('IsNullable') if 'IsNullable' in dmv_row else dmv_row.get('[IsNullable]'))
                            elif ex == 'IsKey':
                                extra_vals.append(dmv_row.get('IsKey') if 'IsKey' in dmv_row else dmv_row.get('[IsKey]'))
                            elif ex == 'SummarizeBy':
                                extra_vals.append(dmv_row.get('SummarizeBy') or dmv_row.get('[SummarizeBy]'))
                            else:
                                extra_vals.append(None)

                    # Pad to 3 for consistent headers
                    while len(values) < 3:
                        values.append(None)
                    core_row = [table_name, col_name, dt]
                    # Insert extras between datatype and sample values
                    core_row.extend(extra_vals)
                    core_row.extend([values[0], values[1], values[2]])
                    result_rows.append(core_row)

            # 4) Write to file (user-selected folder or exports/)
            root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            if output_dir and isinstance(output_dir, str) and output_dir.strip():
                out_dir = output_dir.strip()
            else:
                out_dir = os.path.join(root_dir, "exports")
            try:
                os.makedirs(out_dir, exist_ok=True)
            except Exception:
                # Fallback to default exports folder if user path invalid
                out_dir = os.path.join(root_dir, "exports")
                os.makedirs(out_dir, exist_ok=True)

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fmt = (format or 'csv').strip().lower()
            headers = [
                "Table Name",
                "Column Name",
                "Data Type Column",
            ]
            # Add any requested extras to headers in the same order
            if extras_dmv:
                headers.extend(extras_dmv)
            headers.extend([
                "Sample Row 1 Value",
                "Sample Row 2 Value",
                "Sample Row 3 Value",
            ])

            notes: List[str] = []
            path: str

            # Guard: XLSX can be slow on large exports; auto-switch to CSV for speed
            try:
                if fmt == 'xlsx' and len(result_rows) > 1000:
                    notes.append(f"Result has {len(result_rows)} rows; switched to CSV for faster write")
                    fmt = 'csv'
            except Exception:
                pass

            if fmt == 'txt':
                path = os.path.join(out_dir, f"flat_schema_samples_{ts}.txt")
                with open(path, "w", encoding="utf-8") as f:
                    f.write("\t".join(headers) + "\n")
                    for r in result_rows:
                        f.write("\t".join("" if v is None else str(v) for v in r) + "\n")
            elif fmt == 'xlsx':
                try:
                    from openpyxl import Workbook  # type: ignore
                    wb = Workbook()
                    ws = getattr(wb, 'active', None)
                    if ws is None:
                        raise RuntimeError('openpyxl workbook has no active worksheet')
                    try:
                        ws.title = "SchemaSamples"  # type: ignore[attr-defined]
                    except Exception:
                        pass
                    ws.append(headers)  # type: ignore[attr-defined]
                    for r in result_rows:
                        ws.append(r)  # type: ignore[attr-defined]
                    path = os.path.join(out_dir, f"flat_schema_samples_{ts}.xlsx")
                    wb.save(path)
                except Exception as _e:
                    # Fallback to CSV if openpyxl not available
                    notes.append(f"openpyxl not available or failed ({_e}); fell back to CSV.")
                    fmt = 'csv'
                    path = os.path.join(out_dir, f"flat_schema_samples_{ts}.csv")
                    with open(path, "w", encoding="utf-8", newline="") as f:
                        f.write("sep=,\n")
                        w = csv.writer(f, delimiter=',')
                        w.writerow(headers)
                        w.writerows(result_rows)
            else:
                # csv with Excel-friendly hint (sep=,)
                path = os.path.join(out_dir, f"flat_schema_samples_{ts}.csv")
                with open(path, "w", encoding="utf-8", newline="") as f:
                    # Excel locale hint to ensure comma separation even in locales using ;
                    f.write("sep=,\n")
                    w = csv.writer(f, delimiter=',')
                    w.writerow(headers)
                    w.writerows(result_rows)

            return {
                'success': True,
                'format': fmt,
                'file': path,
                'tables': total_tables,
                'columns': len(result_rows),
                'rows_per_column': n,
                'notes': notes,
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}

    # -------- NL intent executor --------
    def execute_intent(
        self,
        connection_manager,
        connection_state,
        goal: str,
        query: Optional[str] = None,
        table: Optional[str] = None,
        runs: Optional[int] = None,
        max_rows: Optional[int] = None,
        verbose: bool = False,
        candidate_a: Optional[str] = None,
        candidate_b: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Simple rule-based intent executor so users don't need to say 'run safely'.
        Decides which orchestrations to call based on keywords and inputs.
        """
        # Ensure connection first
        ensured = self.ensure_connected(connection_manager, connection_state)
        actions: List[Dict[str, Any]] = [{"action": "ensure_connected", "result": ensured}]
        if not ensured.get("success"):
            return {"success": False, "error": "Connection not established", "phase": "ensure_connected", "details": ensured}

        text = (goal or "").lower()

        # Schema + samples intent: produce flat list with sample values per column
        if any(k in text for k in [
            "sample row", "sample rows", "sample values", "columns with sample", "columns and samples",
            "flat schema", "schema samples", "all tables and columns with sample"
        ]):
            # Determine preferred format from intent
            fmt = 'csv'
            if 'excel' in text or 'xlsx' in text:
                fmt = 'xlsx'
            elif 'txt' in text or 'text' in text:
                fmt = 'txt'
            # Detect desired row count (1-10) if present
            import re
            m = re.search(r"(\d+)\s*(sample|row|rows)", text)
            rows = 3
            try:
                if m:
                    rows = max(1, min(10, int(m.group(1))))
            except Exception:
                rows = 3
            res = self.export_flat_schema_samples(connection_state, fmt, rows)
            return {"success": res.get("success", False), "actions": [{"action": "export_flat_schema_samples", "result": res}], "final": res}

        # Compact/expression-free schema intent
        if any(k in text for k in [
            "compact schema", "expression-free schema", "schema only", "diff-friendly schema", "structure only"
        ]):
            inc_hidden = not ("visible only" in text or "hide hidden" in text)
            cmp_res = self.export_compact_schema(connection_state, include_hidden=inc_hidden)
            actions.append({"action": "export_compact_schema", "result": cmp_res})
            return {"success": cmp_res.get("success", False), "actions": actions, "final": cmp_res}

        # Relationship-focused intents: return list and analysis directly
        if any(k in text for k in ["relationship", "relationships", "rels", "cardinality", "relations"]):
            rel = self.relationship_overview(connection_state)
            actions.append({"action": "relationship_overview", "result": rel})
            return {"success": rel.get("success", False), "actions": actions, "final": rel}

        # Optimization benchmark if both candidates present
        if candidate_a and candidate_b:
            # Use optimize_variants for 2-candidate case to avoid duplicate tooling
            bench = self.optimize_variants(connection_state, [candidate_a, candidate_b], runs)
            actions.append({"action": "optimize_variants", "result": bench})
            return {"success": bench.get("success", False), "actions": actions, "final": bench}

        # Documentation or model summary intents
        if any(k in text for k in ["document", "documentation", "docs", "summarize", "overview", "schema", "structure", "model summary", "list tables", "list measures"]):
            # Prefer a safe summary
            summary = self.summarize_model_safely(connection_state)
            actions.append({"action": "summarize_model", "result": summary})
            if not summary.get("success"):
                return {"success": False, "phase": "summarize_model", "actions": actions, "final": summary}
            # If explicitly asked to document, proceed to docs
            if any(k in text for k in ["document", "documentation", "docs"]):
                doc = self.generate_docs_safe(connection_state)
                actions.append({"action": "generate_docs_safe", "result": doc})
                # If the user also signaled compact schema, attach it as a side artifact
                if any(k in text for k in ["compact", "expression-free", "schema only", "diff-friendly"]):
                    cmp_res = self.export_compact_schema(connection_state)
                    actions.append({"action": "export_compact_schema", "result": cmp_res})
                return {"success": doc.get("success", False), "actions": actions, "final": doc}
            return {"success": True, "actions": actions, "final": summary}

        # Performance analysis intents
        if any(k in text for k in ["analyze performance", "performance", "perf", "se/fe", "storage engine", "formula engine"]):
            if not query:
                return {"success": False, "error": "No query provided for performance analysis", "phase": "input_validation", "actions": actions}
            res = self.safe_run_dax(connection_state, query, mode="analyze", runs=runs, max_rows=max_rows, verbose=verbose)
            actions.append({"action": "safe_run_dax(analyze)", "result": res})
            return {"success": res.get("success", False), "actions": actions, "final": res}

        # If user asks to analyze the model without specifics, propose options (fast vs normal)
        if any(k in text for k in ["analyze", "analysis"]) and not query and not any(k in text for k in ["performance", "perf", "se/fe", "storage engine", "formula engine"]):
            proposal = self.propose_analysis_options(connection_state, goal)
            actions.append({"action": "propose_analysis_options", "result": proposal})
            return {"success": True, "decision": "propose_analysis", "actions": actions, "final": proposal}

        # Generic run/query/preview intents
        if query or any(k in text for k in ["run", "execute", "preview", "query", "show"]):
            res = self.safe_run_dax(connection_state, query or "", mode="preview", runs=runs, max_rows=max_rows, verbose=verbose)
            actions.append({"action": "safe_run_dax(preview)", "result": res})
            return {"success": res.get("success", False), "actions": actions, "final": res}

        # Health/status intents
        if any(k in text for k in ["health", "status", "connected", "ready"]):
            health = self.agent_health(connection_manager, connection_state)
            actions.append({"action": "agent_health", "result": health})
            return {"success": health.get("success", False), "actions": actions, "final": health}

        # Default: plan based on table context
        plan = self.plan_query(text, table, max_rows)
        actions.append({"action": "plan_query", "result": plan})
        return {"success": True, "actions": actions, "final": plan}

    def propose_analysis_options(self, connection_state, goal: Optional[str] = None) -> Dict[str, Any]:
        """Return a simple decision card offering Fast vs Normal analysis using full_analysis.

        - Fast: summary + relationships only (profile=fast, depth=light, include_bpa=false)
        - Normal: full sections (profile=balanced, depth=standard, include_bpa=true when available)
        """
        # Determine if connected for realistic expectation
        connected = connection_state.is_connected()
        notes = []
        if not connected:
            notes.append("Not connected yet; choose an option to auto-connect and run.")
        options: List[Dict[str, Any]] = [
            {
                'name': 'Fast summary',
                'id': 'fast',
                'description': 'Very quick summary of model (tables, columns, measures, relationships) with relationship list.',
                'estimated_time': 'few seconds',
                'call': {
                    'tool': 'full_analysis',
                    'arguments': {
                        'profile': 'fast',
                        'depth': 'light',
                        'include_bpa': False,
                        'limits': {'relationships_max': 200, 'issues_max': 200}
                    }
                }
            },
            {
                'name': 'Normal analysis',
                'id': 'normal',
                'description': 'Comprehensive analysis including best practices and M scan; optionally BPA if available.',
                'estimated_time': 'tens of seconds',
                'call': {
                    'tool': 'full_analysis',
                    'arguments': {
                        'profile': 'balanced',
                        'depth': 'standard',
                        'include_bpa': True,
                        'limits': {'relationships_max': 200, 'issues_max': 200}
                    }
                }
            }
        ]
        # Offer a zero-risk, diff-friendly option
        options.insert(0, {
            'name': 'Compact schema export',
            'id': 'compact',
            'description': 'Export expression-free schema (tables/columns/measures/relationships) for documentation or diffs.',
            'estimated_time': 'seconds',
            'call': {
                'tool': 'export_compact_schema',
                'arguments': {'include_hidden': True}
            }
        })
        return {
            'success': True,
            'decision': 'propose_analysis',
            'goal': goal,
            'connected': connected,
            'options': options,
            'notes': notes
        }
